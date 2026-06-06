"""Выгрузки Excel для админки (openpyxl)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter


def _autosize_columns(ws, max_width: int = 48):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(length + 2, 10)


def _workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(ws, columns: list[tuple[str, str]], rows: list[dict]):
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([row.get(key) for key, _ in columns])
    _autosize_columns(ws)


SUBSCRIBERS_COLUMNS = [
    ("user_id", "ID пользователя"),
    ("username", "Username"),
    ("full_name", "ФИО"),
    ("first_name", "Имя"),
    ("last_name", "Фамилия"),
    ("phone", "Телефон"),
    ("age", "Возраст"),
    ("birth_date", "Дата рождения"),
    ("user_role", "Роль"),
    ("plan", "Тариф"),
    ("paid_until", "Premium до"),
    ("trial_used", "Пробный период использован"),
    ("metro_zones", "Станции метро"),
    ("categories", "Категории"),
    ("registered_at", "Дата регистрации"),
    ("is_active", "Активен"),
    ("has_photo", "Есть фото"),
    ("resume_extra", "Доп. информация"),
]

VACANCIES_COLUMNS = [
    ("id", "ID вакансии"),
    ("category_code", "Категория"),
    ("source_chat_title", "Чат-источник"),
    ("author_contact", "Контакт"),
    ("contact_source", "Источник контакта"),
    ("poster_user_id", "ID автора (TG)"),
    ("poster_username", "Username автора"),
    ("poster_display_name", "Имя автора"),
    ("employer_id", "ID заказчика"),
    ("posted_by_bot_user_id", "ID заказчика в боте"),
    ("address", "Адрес"),
    ("published_at", "Опубликовано"),
    ("found_at", "Найдено парсером"),
    ("is_closed", "Закрыта"),
    ("message_link", "Ссылка"),
    ("message_text", "Текст"),
]

EMPLOYERS_COLUMNS = [
    ("id", "ID"),
    ("telegram_user_id", "Telegram ID"),
    ("username", "Username"),
    ("display_name", "Имя"),
    ("contact_text", "Контакт"),
    ("contact_source", "Источник контакта"),
    ("vacancies_count", "Вакансий"),
    ("categories_csv", "Категории"),
    ("bot_user_id", "ID в боте"),
    ("first_seen_at", "Первый раз"),
    ("last_seen_at", "Последний раз"),
]

NOTFIT_REASON_LABELS = {
    "wrong_category": "Не та категория / роль",
    "low_pay": "Мало платят",
    "wrong_area": "Не мой район / далеко",
    "spam": "Спам или не вакансия",
    "duplicate": "Уже видел / повтор",
    "other": "Другое",
}


def build_subscribers_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Подписчики"
    _write_sheet(ws, SUBSCRIBERS_COLUMNS, rows)
    return _workbook_to_bytes(wb)


def build_vacancies_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Вакансии"
    _write_sheet(ws, VACANCIES_COLUMNS, rows)
    return _workbook_to_bytes(wb)


def build_notfit_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Не подходит"
    columns = [
        ("id", "ID записи"),
        ("created_at", "Дата"),
        ("user_id", "ID пользователя"),
        ("username", "Username"),
        ("full_name", "ФИО"),
        ("reason_code", "Код причины"),
        ("reason_label", "Причина"),
        ("reason_text", "Комментарий"),
        ("vacancy_id", "ID вакансии"),
        ("vacancy_category", "Категория вакансии"),
        ("user_categories", "Категории пользователя"),
        ("source_chat_title", "Чат"),
        ("message_link", "Ссылка"),
        ("message_text", "Текст вакансии"),
    ]
    enriched = []
    for r in rows:
        code = r.get("reason_code") or ""
        enriched.append({
            **r,
            "reason_label": NOTFIT_REASON_LABELS.get(code, code),
            "full_name": r.get("full_name") or r.get("first_name"),
            "vacancy_category": r.get("vacancy_category") or r.get("vacancy_category_live"),
        })
    _write_sheet(ws, columns, enriched)
    return _workbook_to_bytes(wb)


DRAFT_STATUS_EXPORT_LABELS = {
    "delivered": "Черновик готов (кнопка в боте)",
    "manual": "Вручную (без кнопки чата)",
    "failed": "Сбой доставки черновика",
    "pending": "В обработке",
}


def build_responses_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отклики"
    columns = [
        ("id", "ID отклика"),
        ("responded_at", "Дата отклика"),
        ("user_id", "ID пользователя"),
        ("username", "Username"),
        ("full_name", "ФИО"),
        ("phone", "Телефон"),
        ("vacancy_id", "ID вакансии"),
        ("category_code", "Категория"),
        ("source_chat_title", "Чат-источник"),
        ("employer_contact", "Контакт заказчика"),
        ("draft_status", "Код статуса черновика"),
        ("draft_status_label", "Статус черновика"),
        ("response_status", "Статус отклика"),
        ("vacancy_closed", "Вакансия закрыта"),
        ("star_boost", "Расширенный отклик (Stars)"),
        ("vacancy_link", "Ссылка на пост"),
        ("vacancy_text", "Текст вакансии (фрагмент)"),
    ]
    enriched = []
    for r in rows:
        code = r.get("draft_status") or "pending"
        enriched.append({
            **r,
            "draft_status_label": DRAFT_STATUS_EXPORT_LABELS.get(code, code),
            "vacancy_closed": "да" if r.get("vacancy_closed") else "нет",
            "star_boost": "да" if r.get("star_boost") else "нет",
        })
    _write_sheet(ws, columns, enriched)
    return _workbook_to_bytes(wb)


def build_employers_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказчики"
    _write_sheet(ws, EMPLOYERS_COLUMNS, rows)
    return _workbook_to_bytes(wb)


def export_filename(prefix: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{prefix}_{stamp}.xlsx"
